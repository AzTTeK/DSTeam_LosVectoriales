import openpyxl
import os
import auth

class Paquetes:
    def __init__(self, nombre, peso, precio, tipo, contenido=None, categoria=None, dimension=None, estado_pedido=None,creador = auth.usuario, direccion = None, domiciliario = None):
        self.__nombre = nombre
        self.__peso = f"{peso} kg"
        self.__precio = f"{precio} $"
        self.__tipo = tipo
        self.__contenido = contenido if contenido else []
        self.__categoria = categoria if categoria else []
        self.__dimension = dimension if dimension else []
        self.__estado_pedido = estado_pedido
        self.__creador = auth.usuario
        self.__direccion = direccion
        self.__domiciliario = domiciliario


    @property
    def nombre(self):
        return self.__nombre
    
    @nombre.setter
    def nombre(self, new_name):
        self.__nombre = new_name

    @property
    def peso(self):
        return self.__peso
    
    @peso.setter
    def peso(self, new_peso):
        self.__peso = f"{new_peso} kg"

    @property
    def tipo(self):
        return self.__tipo

    @property
    def precio(self):
        return self.__precio

    @precio.setter
    def precio(self, new_price):
        self.__peso = new_price
    
    @tipo.setter
    def tipo(self, new_tipo):
        self.__tipo = new_tipo

    @property
    def contenido(self):
        return self.__contenido
    
    @contenido.setter
    def contenido(self, cont):
        self.__contenido.pop()
        self.__contenido.append(cont)

    @property
    def categoria(self):
        return self.__categoria
    
    @categoria.setter
    def categoria(self, cat):
        self.__categoria.pop()
        self.__categoria.append(cat)

    @property
    def dimension(self):
        return self.__dimension
    
    @dimension.setter
    def dimension(self, dim):
        self.__dimension.pop()
        self.__dimension.append(dim)

    @property
    def estado_pedido(self):
        return self.__estado_pedido
    
    def agregar_estado_pedido(self, nuevo_estado):
        self.__estado_pedido = nuevo_estado

    @property
    def direccion(self):
        return self.__direccion

    @direccion.setter
    def direccion(self, new_dir):
        self.__direccion = new_dir

    @property
    def domiciliario(self):
        return self.__domiciliario

    @domiciliario.setter
    def precio(self, new_dom):
        self.__domiciliario = new_dom

    def __str__(self):
        return (f"Creador: {self.__creador}\n"
                f"Paquete: {self.__nombre}\n"
                f"precio: {self.__precio}\n"
                f"Peso: {self.__peso}\n"
                f"Tipo: {self.__tipo}\n"
                f"Contenido: {', '.join(self.__contenido) if self.__contenido else 'N/A'}\n"
                f"Categoría: {', '.join(self.__categoria) if self.__categoria else 'N/A'}\n"
                f"Dimensiones: {', '.join(self.__dimension) if self.__dimension else 'N/A'}\n"
                f"estado del pedido: {self.__estado_pedido}\n")
                
class Creacion:
    @staticmethod
    def crear_paquete():

        nombre = input("Nombre del paquete: ")
        while nombre in column[0]:
            print("Nombre de paquete ya existente...")
            nombre = input("Nombre del paquete: ")
        peso = input("Peso (kg): ")
        tipo = None
        while tipo != "basico" or tipo != "estandar" or tipo != "dimensionado":
            print("El tipo debe ser basico, estandar o dimensionado") 
            tipo = input("Tipo (basico, estandar o dimensionado): ")
            if tipo == "basico" or tipo == "estandar" or tipo == "dimensionado":
                break
        contenido = input("Contenido (separado por comas): ").split(",") 
        categoria = input("Categorías (separadas por comas): ").split(",") 
        dimension = input("Dimensiones (ej: 10x20x30 cm): ").split(",")
        estado_pedido = "pendiente"
        return Paquetes(nombre, peso, tipo, contenido, categoria, dimension, estado_pedido)

class ExcelPaquetes:
    FILE_PATH = "paquetes.xlsx"

    @classmethod
    def iniciar_excel(cls):
        if not os.path.exists(cls.FILE_PATH):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Creador", "Nombre", "Precio", "Peso", "Tipo", "Contenido", "Categoría", "Dimensiones", "estado pedido","Direccion","Domiciliario"])
            wb.save(cls.FILE_PATH)

    @classmethod
    def guardar(cls, paquete):
        if not os.path.exists(cls.FILE_PATH):
            cls.iniciar_excel()

        wb = openpyxl.load_workbook(cls.FILE_PATH)
        ws = wb.active
        ws.append([
            paquete.creador,
            paquete.nombre,
            paquete.precio,
            paquete.peso,
            paquete.tipo,
            ", ".join(paquete.contenido) if paquete.contenido else "N/A",
            ", ".join(paquete.categoria) if paquete.categoria else "N/A",
            ", ".join(paquete.dimension) if paquete.dimension else "N/A",
            paquete.estado_pedido,
            paquete.direccion,
            paquete.domiciliario,
            ])
        
        wb.save(cls.FILE_PATH)
        wb.close()
        print("\nPaquete guardado con éxito en la base de datos.")

    @classmethod
    def mostrar(cls):
        if not os.path.exists(cls.FILE_PATH):
            print("No hay paquetes guardados.")
            return

        wb = openpyxl.load_workbook(cls.FILE_PATH)
        ws = wb.active
        
        if ws.max_row == 1:
            print("No hay paquetes almacenados.")
            return

        print("\nPaquetes almacenados:")
        for row in ws.iter_rows(min_row=2, values_only=True):
            print(f"{row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]} | {row[5]} | {row[6]} | {row[7]} | {row[8]}" )
         
        wb.close()
    @classmethod    
    def mostrar_dis(cls):
        if not os.path.exists(cls.FILE_PATH):
            print("No hay paquetes guardados.")
            return

        wb = openpyxl.load_workbook(cls.FILE_PATH)
        ws = wb.active
        
        if ws.max_row == 1:
            print("No hay paquetes almacenados.")
            return

        print("\nPaquetes almacenados:")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[8] == "pendiente":
                print(f"{row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]} | {row[5]} | {row[6]} | {row[7]} | Disponible ")
         
        wb.close()
    @classmethod
    def actualizar_datos(cls):
        paquetes_usuario =
        if usuario in row[0]:
            print("Tus productos: ")
            j = 1
            for i in  ws.iter_rows(min_row=2, values_only=True):
                if i[0]==usuario:
                    print(f"{j} | {row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]} | {row[5]} | {row[6]} | {row[7]} ")
                    j+=1
            pcambio=int(input("Que paquete desea cambiar? "))
            j = 1
            for i in  ws.iter_rows(min_row=2, values_only=True):
                if i[0] == usuario:
                    j+=1
                if i[0] == usuario and j == pcambio:
                    print(f"1.{row[1]} | 2.{row[2]} | 3.{row[3]} | 4.{row[4]} | 5.{row[5]} | 6.{row[6]} | 7.{row[7]} ")
                    cambio = int(input("Digite el numero que esta al lado de lo que quiere cambiar"))
                    if cambio==1:
                        nuevon = input("Digite el nuevo nombre")
                        nombre(nuevon)
                        print("Cambio realizado")
                    if cambio==2:
                        nuevopr = input("Digite el nuevo precio")
                        precio(nuevop)
                        print("Cambio realizado")
                    if cambio==3:
                        nuevope = input("Digite el nuevo peso")
                        peso(nuevope)
                        print("Cambio realizado")
                    if cambio==4:
                        nuevot = input("Digite el nuevo tipo")
                        tipo(nuevot)
                        print("Cambio realizado")
                    if cambio==5:
                        nuevoco = input("Digite el nuevo contenido")
                        contenido(nuevoco)
                        print("Cambio realizado")
                    if cambio==6:
                        nuevonca = input("Digite el nuevo categoria")
                        categoria(nuevonca)
                        print("Cambio realizado")
                    if cambio==7:
                        nuevod = input("Digite el nuevo dimension")
                        dimension(nuevod)
                        print("Cambio realizado")


    @classmethod
    def registrar_pedido(cls):
        pass


if __name__ == "main":
    paquete = Creacion.crear_paquete()
    print("\nPaquete creado con éxito:\n")
    print(paquete)

    ExcelPaquetes.iniciar_excel()
    ExcelPaquetes.guardar(paquete)
    ExcelPaquetes.mostrar()
    ExcelPaquetes.mostrar_dis()